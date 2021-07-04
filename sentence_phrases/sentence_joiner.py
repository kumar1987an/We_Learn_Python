# Loading Libraries
import openpyxl as pyxl
import json

# Function definition for join_sentence framework


def join_sentence(filename):
    # File loading into data frame
    df = pyxl.load_workbook(filename)

    # Loading the active worksheet
    sheet = df.active

    # Parsing first phrase column C1 strings
    first_phrase_strings = [
        c1[0]
        for c1 in sheet.iter_rows(
            min_row=2, max_row=19, min_col=2, max_col=2, values_only=True
        )
    ]

    # Parsing second phrase column C2 strings
    second_phrase_strings = [
        c2[0]
        for c2 in sheet.iter_rows(
            min_row=2, max_row=19, min_col=3, max_col=3, values_only=True
        )
        if c2[0] != None
    ]

    json_data = {}  # empty dictionary for loading resulted sentence
    sentence_id = 0  # instantiator for sentence id

    # Loop to bond the sentence and collect required data

    for first_phrase in first_phrase_strings:
        for second_phrase in second_phrase_strings:
            if second_phrase != "None":
                sentence_id = sentence_id + 1
                sentence_bond = first_phrase + " " + second_phrase
                entries = {
                    "text": sentence_bond,
                    "end": len(sentence_bond),
                    "start": len(first_phrase) + 1,
                    "value": second_phrase,
                }
                json_data[f"SentenceID_{sentence_id}"] = entries

    # Writing  json_data as a json file "sentences_bond.json"
    with open("json_data.json", "w") as file:
        file.write(json.dumps(json_data))


# Function call for joining sentence
join_sentence("data_parse.xlsx")
